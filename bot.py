import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
import yadisk
import tempfile
import asyncio

from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

# ================= НАСТРОЙКИ =================
TOKEN = os.environ.get("TOKEN")
YANDEX_TOKEN = os.environ.get("YANDEX_TOKEN")
WEBHOOK_URL = os.environ.get("WEBHOOK_URL")  # https://<название-приложения>.up.railway.app/bot

if not TOKEN:
    raise Exception("❌ Telegram TOKEN не задан в переменных окружения")
if not YANDEX_TOKEN:
    raise Exception("❌ YANDEX_TOKEN не задан в переменных окружения")
if not WEBHOOK_URL:
    raise Exception("❌ WEBHOOK_URL не задан в переменных окружения")

SUPERUSERS = {805289423, 502894278}
DISK_FOLDER = "/SberBot"
ALLOWED_FILE = "allowed_users.json"

# ================= YANDEX DISK =================
y = yadisk.YaDisk(token=YANDEX_TOKEN)
if not y.check_token():
    raise Exception("❌ Yandex token недействителен")
if not y.exists(DISK_FOLDER):
    y.mkdir(DISK_FOLDER)

def disk_path(filename):
    return f"{DISK_FOLDER}/{filename}"

# ================= ВРЕМЕННЫЕ ФАЙЛЫ =================
def temp_path(filename):
    return os.path.join(tempfile.gettempdir(), filename)

def cleanup_temp(*files):
    for f in files:
        if f and os.path.exists(f):
            os.remove(f)

# ================= ФАЙЛЫ НА ДИСКЕ =================
def download_file(filename):
    local_file = temp_path(filename)
    if y.exists(disk_path(filename)):
        y.download(disk_path(filename), local_file)
        return local_file
    return None

def upload_file(filename):
    local_file = temp_path(filename)
    y.upload(local_file, disk_path(filename), overwrite=True)

# ================= ДОСТУП =================
def load_allowed():
    local_file = download_file(ALLOWED_FILE)
    if not local_file:
        temp_file = temp_path(ALLOWED_FILE)
        with open(temp_file, "w", encoding="utf-8") as f:
            json.dump(list(SUPERUSERS), f)
        upload_file(ALLOWED_FILE)
        cleanup_temp(temp_file)
        return set(SUPERUSERS)
    with open(local_file, "r", encoding="utf-8") as f:
        users = set(json.load(f))
    cleanup_temp(local_file)
    return users.union(SUPERUSERS)

def save_allowed(users):
    temp_file = temp_path(ALLOWED_FILE)
    with open(temp_file, "w", encoding="utf-8") as f:
        json.dump(list(users), f)
    upload_file(ALLOWED_FILE)
    cleanup_temp(temp_file)

ALLOWED_USERS = load_allowed()

def has_access(user_id):
    return user_id in SUPERUSERS or user_id in ALLOWED_USERS

# ================= КЛАВИАТУРА =================
def main_keyboard(user_id):
    buttons = []
    if has_access(user_id):
        buttons += [
            ["📖 Показать записи", "📥 Скачать Excel"],
            ["🧹 Очистить файл", "❌ Удалить строку"],
            ["🗂 Архив Excel"]
        ]
    if user_id in SUPERUSERS:
        buttons += [["👑 Управление доступом"]]
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True, is_persistent=True)

# ================= EXCEL =================
def get_today_filename():
    return f"data_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

def ensure_file(filename):
    local_file = download_file(filename)
    if not local_file:
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "ВСП", "ИНН", "Наименование", "Бумага/эл", "User"])
        wb.save(temp_path(filename))
        upload_file(filename)
        cleanup_temp(temp_path(filename))

def append_row(filename, row):
    local_file = download_file(filename)
    if not local_file:
        ensure_file(filename)
        local_file = temp_path(filename)
    wb = load_workbook(local_file)
    ws = wb.active
    ws.append(row)
    count = ws.max_row - 1
    wb.save(local_file)
    upload_file(filename)
    cleanup_temp(local_file)
    return count

def get_rows(filename):
    local_file = download_file(filename)
    if not local_file:
        return []
    wb = load_workbook(local_file)
    ws = wb.active
    rows = [
        f"{i+1}. {' | '.join(map(str, r[1:5]))}"
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True))
    ]
    cleanup_temp(local_file)
    return rows

def delete_row(filename, idx):
    local_file = download_file(filename)
    wb = load_workbook(local_file)
    ws = wb.active
    ws.delete_rows(idx + 1)
    wb.save(local_file)
    upload_file(filename)
    cleanup_temp(local_file)

def clear_file(filename):
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата", "ВСП", "ИНН", "Наименование", "Бумага/эл", "User"])
    wb.save(temp_path(filename))
    upload_file(filename)
    cleanup_temp(temp_path(filename))

def list_excel_files():
    items = y.listdir(DISK_FOLDER)
    return [i["name"] for i in items if i["type"] == "file" and i["name"].endswith(".xlsx")]

# ================= СОСТОЯНИЯ =================
WAITING_DELETE = set()
WAITING_CLEAR_CONFIRM = set()
WAITING_ARCHIVE_SELECT = dict()

# ================= БОТ =================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    await update.message.reply_text(
        "🤖 Бот учёта сообщений.\n\n"
        "Отправь сообщение из 4 строк:\n"
        "1 — ВСП\n2 — ИНН\n3 — Наименование\n4 — Бумага/эл",
        reply_markup=main_keyboard(user_id),
    )

async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_id = update.message.from_user.id

    if not has_access(user_id):
        await update.message.reply_text("❌ Нет доступа.")
        return

    filename = get_today_filename()
    ensure_file(filename)

    # --- Админ ---
    if user_id in SUPERUSERS:
        if text == "👑 Управление доступом":
            await update.message.reply_text("+ ID — дать доступ\n- ID — забрать доступ")
            return
        if text.startswith("+"):
            uid = int(text[1:].strip())
            ALLOWED_USERS.add(uid)
            save_allowed(ALLOWED_USERS)
            await update.message.reply_text(f"Доступ выдан: {uid}")
            return
        if text.startswith("-"):
            uid = int(text[1:].strip())
            ALLOWED_USERS.discard(uid)
            save_allowed(ALLOWED_USERS)
            await update.message.reply_text(f"Доступ забран: {uid}")
            return

    # --- Кнопки ---
    if text == "📖 Показать записи":
        rows = get_rows(filename)
        msg = "\n".join(rows) if rows else "Нет записей."
        await update.message.reply_text(msg, reply_markup=main_keyboard(user_id))
        return

    if text == "📥 Скачать Excel":
        local_file = download_file(filename)
        if local_file:
            await update.message.reply_document(open(local_file, "rb"), reply_markup=main_keyboard(user_id))
            cleanup_temp(local_file)
        return

    if text == "🧹 Очистить файл":
        WAITING_CLEAR_CONFIRM.add(user_id)
        await update.message.reply_text("Напишите ДА для подтверждения.")
        return
    if user_id in WAITING_CLEAR_CONFIRM:
        if text.upper() == "ДА":
            clear_file(filename)
            await update.message.reply_text("Файл очищен.", reply_markup=main_keyboard(user_id))
        else:
            await update.message.reply_text("Файл не был очищен.", reply_markup=main_keyboard(user_id))
        WAITING_CLEAR_CONFIRM.discard(user_id)
        return

    if text == "❌ Удалить строку":
        WAITING_DELETE.add(user_id)
        await update.message.reply_text("Введите номер строки:")
        return
    if user_id in WAITING_DELETE:
        try:
            idx = int(text)
            delete_row(filename, idx)
            await update.message.reply_text(f"Удалена строка {idx}.", reply_markup=main_keyboard(user_id))
        except:
            await update.message.reply_text("Введите корректное число.", reply_markup=main_keyboard(user_id))
        WAITING_DELETE.discard(user_id)
        return

    if text == "🗂 Архив Excel":
        files = list_excel_files()
        if not files:
            await update.message.reply_text("Архив пуст.", reply_markup=main_keyboard(user_id))
            return
        WAITING_ARCHIVE_SELECT[user_id] = files
        msg = "\n".join([f"{i+1}. {f}" for i, f in enumerate(files)])
        await update.message.reply_text(f"Выберите файл для скачивания по номеру:\n{msg}")
        return
    if user_id in WAITING_ARCHIVE_SELECT:
        try:
            idx = int(text) - 1
            files = WAITING_ARCHIVE_SELECT[user_id]
            if 0 <= idx < len(files):
                local_file = download_file(files[idx])
                await update.message.reply_document(open(local_file, "rb"), reply_markup=main_keyboard(user_id))
                cleanup_temp(local_file)
            else:
                await update.message.reply_text("Некорректный номер.", reply_markup=main_keyboard(user_id))
        except:
            await update.message.reply_text("Введите число.", reply_markup=main_keyboard(user_id))
        WAITING_ARCHIVE_SELECT.pop(user_id)
        return

    # --- Добавление записи ---
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if len(lines) != 4:
        await update.message.reply_text(f"❌ Нужно 4 строки, получено {len(lines)}.", reply_markup=main_keyboard(user_id))
        return

    username = update.message.from_user.username or update.message.from_user.full_name
    count = append_row(filename, [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), *lines, username])
    await update.message.reply_text(f"Добавлено. Всего строк: {count}", reply_markup=main_keyboard(user_id))

# ================= ЗАПУСК =================
async def main():
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))

    # Запуск webhook на Railway
    app.run_webhook(
        listen="0.0.0.0",
        port=int(os.environ.get("PORT", 3000)),
        webhook_url=WEBHOOK_URL
    )

if __name__ == "__main__":
    asyncio.run(main())
